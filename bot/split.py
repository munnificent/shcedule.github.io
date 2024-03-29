import openpyxl

def split_and_process_excel_file(filename):
    # Загрузка исходного файла Excel
    original_wb = openpyxl.load_workbook(filename)
    
    for sheet_name in original_wb.sheetnames:
        # Создание новой книги и копирование содержимого листа
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name
        
        original_ws = original_wb[sheet_name]
        for row in original_ws.iter_rows(values_only=True):
            new_ws.append(row)
        
        # Удаление первых 15 строк
        remove_first_rows(new_ws, 15)
        
        # Заполнение промежутков между днями недели
        fill_days_in_column(new_ws)
        
        # Сохранение новой книги
        new_filename = f"{sheet_name}.xlsx"
        new_wb.save(new_filename)
        print(f"Лист {sheet_name} сохранен в файл {new_filename}")

def remove_first_rows(ws, rows_to_delete):
    for _ in range(rows_to_delete):
        ws.delete_rows(1)

def fill_days_in_column(ws):
    current_day = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value is not None and isinstance(cell.value, str):
            current_day = cell.value
        else:
            cell.value = current_day

# Пример использования:
split_and_process_excel_file("schedule.xlsx")
