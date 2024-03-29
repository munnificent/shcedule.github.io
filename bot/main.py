import telebot
from telebot import types
from openpyxl import load_workbook
import pendulum
import sqlite3
from PIL import Image, ImageDraw, ImageFont
import io

API_TOKEN = '6108466885:AAGXFfFy9Wbz8_T72M3g8QAuqCH_bZOIBB4'
bot = telebot.TeleBot(API_TOKEN)
subject_abbreviations = {
    'Дифференциальное исчисление функций одной переменной': 'Дифф. исчисление',
    'Информационно-коммуникационные технологии': 'ИКТ',
    'Ақпараттық-коммуникациялық технологи':'Ақпараттық тех',
    'Линейная алгебра и аналитическая геометрия': 'Линал',
    'Балалардың жас және физиологиялық даму ерекшеліктері': 'физиология',
    'Возрастные и физиологические особенности развития детей': 'физиология',
    'Элементарная математика (геометрия)': 'геометрия',
    'Элементарлық математика (геометрия)': 'геометрия',
    'Алгоритмы и структуры данных ': 'Алгоритмы',
    'Алгоритмдер және мәліметтер құрылымы': 'Алгоритмдер',
    'Математический анализ-2': 'Матан 2',
    'Математикалық талдау-2': 'Матан 2'

    # Add other subjects and abbreviations here
}

def init_db():
    with sqlite3.connect('schedule.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS users
                     (id INTEGER PRIMARY KEY, first_name TEXT, last_name TEXT, username TEXT, 
                     user_group TEXT, language TEXT DEFAULT 'ru')''')
        conn.commit()
init_db()
@bot.message_handler(commands=['start', 'register'])
def send_welcome(message):
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    markup.add('RU', 'KZ')  # Russian and Kazakh language options
    msg = bot.send_message(
        message.chat.id, 
        "Добро пожаловать в бота управления расписанием!\n"
        "Выберите язык / Таңдаңыз тіліңіз:",
        reply_markup=markup
    )
    bot.register_next_step_handler(msg, process_language_selection)

def process_language_selection(message):
    language = message.text.lower()
    if language not in ['ru', 'kz']:
        msg = bot.send_message(message.chat.id, "Пожалуйста, выберите язык, используя кнопки ниже.")
        bot.register_next_step_handler(msg, process_language_selection)
        return
    
    msg = bot.send_message(message.chat.id, f"Отлично! Выбран язык: {language.upper()}.\nТеперь отправьте вашу группу. Например, 'МОР231'.")
    bot.register_next_step_handler(msg, process_group_registration, language)

def process_group_registration(message, language):
    user_group = message.text.strip().upper()
    if not user_group:
        msg = bot.send_message(message.chat.id, "Вы не указали группу. Пожалуйста, отправьте вашу группу.")
        bot.register_next_step_handler(msg, process_group_registration, language)
        return
    
    user_id = message.from_user.id
    register_user(user_id, user_group, language)
    bot.send_message(
        message.chat.id, 
        "Спасибо за регистрацию! Теперь вы можете использовать команды.",
        reply_markup=types.ReplyKeyboardRemove()
    )


def register_user(user_id, user_group, language, first_name=None, last_name=None, username=None):
    with sqlite3.connect('schedule.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''INSERT INTO users (id, first_name, last_name, username, user_group, language) 
                          VALUES (?, ?, ?, ?, ?, ?) ON CONFLICT(id) DO UPDATE SET 
                          first_name=excluded.first_name, last_name=excluded.last_name, 
                          username=excluded.username, user_group=excluded.user_group, language=excluded.language''',
                       (user_id, first_name, last_name, username, user_group, language))
        conn.commit()

def get_user_group_and_language(user_id):
    with sqlite3.connect('schedule.db') as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT user_group, language FROM users WHERE id = ?', (user_id,))
        return cursor.fetchone() or (None, None)

# Bot command handlers

# ... include previous command handlers (/start, /register, /re-register, /delete-account, /help, /admins)
@bot.message_handler(commands=['reregister'])
def re_register(message):
    msg = bot.reply_to(message, "Введите вашу новую группу:")
    bot.register_next_step_handler(msg, process_re_registration_step)

def process_re_registration_step(message):
    user_group = message.text.strip()
    user_id = message.from_user.id
    conn = sqlite3.connect('schedule.db')
    cursor = conn.cursor()
    cursor.execute('UPDATE users SET user_group = ? WHERE id = ?', (user_group, user_id))
    conn.commit()
    conn.close()
    bot.reply_to(message, "Ваша информация обновлена.")

@bot.message_handler(commands=['deleteaccount'])
def delete_account(message):
    user_id = message.from_user.id
    conn = sqlite3.connect('schedule.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    bot.reply_to(message, "Ваш аккаунт удален.")

@bot.message_handler(commands=['help'])
def send_help(message):
    help_text = """
    Доступные команды:
    /help - получить список доступных команд
    /reregister - перерегистрировать пользователя
    /delete-account - удалить аккаунт пользователя
    /schedule - получить расписание на неделю
    /schedulenow - получить расписание на текущий день
    /admins - получить контакт администратора
    """
    bot.reply_to(message, help_text)

@bot.message_handler(commands=['admins'])
def send_admin_contact(message):
    bot.reply_to(message, "Контакт администратора: @munificent_archon")
# Helper functions

def create_schedule_image(schedule, title, language):
    width = 800
    height = max(60, len(schedule) * 30 + 60)
    image = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(image)
    
    try:
        font = ImageFont.truetype("arial.ttf", size=20)
    except IOError:
        font = ImageFont.load_default()
    
    draw.text((10, 10), title, fill="black", font=font)
    
    for i, entry in enumerate(schedule):
        try:
            # Ensure all entries are strings and default to empty string if not found
            time = str(entry.get('time', ''))
            subject = subject_abbreviations.get(entry.get('subject', ''), entry.get('subject', ''))
            room = str(entry.get('room', ''))
            teacher = str(entry.get('teacher', ''))
            class_type = str(entry.get('type', '')).lower()
            class_type_text = 'Лекция' if 'лек' in class_type else 'Семинар'
            
            entry_text = f"{time}: {subject} - {room}, {teacher} ({class_type_text})"
            draw.text((10, 40 + i * 30), entry_text, fill="black", font=font)
        except KeyError as e:
            print(f"Missing key in entry: {e}")
            continue  # Skip this entry or log it as needed

    img_byte_arr = io.BytesIO()
    image.save(img_byte_arr, format='PNG')
    img_byte_arr.seek(0)
    
    return img_byte_arr


def get_schedule_from_file(user_group, language, day=None):
    workbook = load_workbook(f'{user_group}.xlsx')
    sheet = workbook.active
    schedule = []
    current_day = pendulum.now().format('dddd', locale='ru').capitalize() if day is None else day

    for row in sheet.iter_rows(values_only=True):
        if row[0] and current_day in row[0]:
            try:
                subject = row[2] if isinstance(row[2], str) else ""
                # Handle subject based on language
                subject_parts = subject.split('/')
                subject = subject_parts[0].strip() if language == 'kz' else subject_parts[1].strip() if len(subject_parts) > 1 else subject_parts[0].strip()

                # Abbreviate subject if applicable
                subject = subject_abbreviations.get(subject, subject)

                # Safely access other row elements
                time = str(row[1]) if row[1] else ""
                type_of_class = str(row[3]) if row[3] else ""
                teacher_name = str(row[4]) if row[4] else ""
                classroom = str(row[5]) if row[5] else ""

                # Handle teacher name based on language
                teacher_parts = teacher_name.split('/')
                teacher = teacher_parts[0].strip() if language == 'kz' else teacher_parts[1].strip() if len(teacher_parts) > 1 else teacher_parts[0].strip()

                schedule.append({
                    'time': time,
                    'subject': subject,
                    'type': type_of_class,
                    'teacher': teacher,
                    'room': classroom
                })
            except IndexError as e:
                # Log or print the error for debugging
                print(f"Error processing row: {row} - {e}")
    
    return schedule




def send_schedule_message(chat_id, user_group, language, day=None):
    schedule = get_schedule_from_file(user_group, language, day)
    if schedule:
        if day:
            title = f"Расписание на {day}" if language == 'ru' else f"{day} күнгі сабақ кестесі"
        else:
            current_day = pendulum.now().format('dddd', locale='ru').capitalize()
            title = "Расписание на сегодня" if language == 'ru' else "Бүгінгі сабақ кестесі"
        img = create_schedule_image(schedule, title, language)
        bot.send_photo(chat_id, photo=img)
    else:
        no_classes_message = "На сегодня занятий нет." if language == 'ru' else "Бүгін сабақтар жоқ."
        bot.send_message(chat_id, no_classes_message)

@bot.message_handler(commands=['schedule'])
def send_weekly_schedule(message):
    user_group, language = get_user_group_and_language(message.from_user.id)
    if not user_group:
        bot.reply_to(message, "Вы не зарегистрированы. Пожалуйста, используйте команду /start для регистрации.")
        return
    
    days_of_week_ru = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    days_of_week_kz = ["Дүйсенбі", "Сейсенбі", "Сәрсенбі", "Бейсенбі", "Жұма", "Сенбі"]
    days_of_week = days_of_week_kz if language == 'kz' else days_of_week_ru
    
    for day in days_of_week:
        send_schedule_message(message.chat.id, user_group, language, day)

@bot.message_handler(commands=['schedulenow'])
def send_today_schedule(message):
    user_group, language = get_user_group_and_language(message.from_user.id)
    if not user_group:
        bot.reply_to(message, "Вы не зарегистрированы. Пожалуйста, используйте команду /start для регистрации.")
        return
    
    send_schedule_message(message.chat.id, user_group, language)

# ... (rest of the bot code remains unchanged)

if __name__ == '__main__':
    bot.polling(none_stop=True)
